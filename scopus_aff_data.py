import pandas as pd
import requests
import json

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    # create a writer for this month and year
    writer = pd.ExcelWriter(filename, engine='openpyxl')

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if not startrow and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # copy existing sheets
        writer.sheets = dict(
            (ws.title, ws) for ws in writer.book.worksheets)
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if not startrow:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


#import Scimago rankings, clean file
df = pd.read_csv('scim173.csv', skip_blank_lines = True, delimiter = ';')
df.columns = df.columns.str.replace(',', '')
df.columns = df.columns.str.replace(' ', '_')
df['Institution'] = df['Institution'].map(lambda x: x.rstrip('*').strip())
df['Sector'] = df['Sector'].map(lambda x: x.rstrip(',').strip())

myKey = #type in your api key
headers = {'accept':'application/json', 'x-els-apikey':myKey}
url = 'http://api.elsevier.com/content/search/affiliation?query=AFFIL%28'
dataFile = 'C:/Users/vkk/desktop/affiliation_data.xlsx'

for row in df.itertuples():
	inst = row.Institution.replace(' ', ' AND ')
	index = row.Index
	resp = requests.get(url + inst + '%29', headers = headers)
	if resp.status_code == 200:
		afDat = resp.json()
		dfr = pd.DataFrame(afDat['search-results']['entry'])
		rank = pd.DataFrame(df.iloc[[index]].values.repeat(len(dfr), axis = 0))
		afDatRank = pd.concat([dfr, rank], axis = 1)
		append_df_to_excel(dataFile, afDatRank)
	else:
		pass


