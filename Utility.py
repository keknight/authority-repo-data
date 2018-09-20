##Utility file with helper functions

import json
import pandas as pd
from pandas_datareader import wb


#helper function to append json to file as valid json
def append_to_json(_dict,path): 
	with open(path, 'ab+') as f:
		f.seek(0,2)
		if f.tell() == 0:
			f.write(json.dumps([_dict]).encode())
		else:
			f.seek(-1,2)
			f.truncate()
			f.write(' , '.encode())
			f.write(json.dumps(_dict).encode())
			f.write(']'.encode())
			

#function to flatten json output 
def flatten_json(y):
	out = {}
	def flatten(x, name = ''):
		if type(x) is dict:
			for a in x:
				flatten(x[a], name + a + '_')
		elif type(x) is list:
			i = 0
			for a in x:
				flatten(a, name)
				i += 1
		else:
			out[name[:-1]] = x
	flatten(y)
	return out
	
			
#creates a number generator
#change num #= to increment by a different amount
def firstn(n):
	num = 0
	while num < n:
		yield num
		num += 25

		
#function to append dataframe to Excel file
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None, **to_excel_kwargs):

	from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
	if 'engine' in to_excel_kwargs:
		to_excel_kwargs.pop('engine')

	writer = pd.ExcelWriter(filename, engine='openpyxl')

	try:
        # try to open an existing workbook
		writer.book = load_workbook(filename)
        # get the last row in the existing Excel sheet
		if not startrow and sheet_name in writer.book.sheetnames:
			startrow = writer.book[sheet_name].max_row

        # copy existing sheets
		writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
	except FileNotFoundError:
        # file does not exist yet
		pass

	if not startrow:
		startrow = 0

	df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
	writer.save()
	
#get country data from pandas_datareader (World Bank data)
#writes directly to Excel file
def get_country_data():
	country_data = wb.get_countries().fillna('none')
	writer = pd.ExcelWriter('World_Bank_country_data.xlsx')
	country_data.to_excel(writer, 'Country Data')
	writer.save()